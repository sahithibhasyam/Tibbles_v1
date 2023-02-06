from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from .models import *
import mysql.connector as sql
from django.db import *
from . import allot
from django.views.generic import View
from django.http import HttpResponse
from .utils import render_to_pdf
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ------------------------------------------------------------------------------

def myhome(request):
    return render(request, "mainpage.html")

#------------------------------------------------------------------------------

def about(request):
    return render(request, "about.html")

#------------------------------------------------------------------------------

def home(request):
    if request.user.is_authenticated:
        if request.user.username == '11':
            cursor = connection.cursor()
            cursor.execute("SELECT DISTINCT f.faculty_id, f.faculty_name FROM faculty F NATURAL JOIN pref_sub p NATURAL JOIN pref_slot s WHERE f.faculty_id = p.faculty_id AND f.faculty_id = s.faculty_id order by f.faculty_id;")
            fp = cursor.fetchall()
            return render(request, 'home.html', {'fp1': fp} )
        else:
            return render(request, "dashboard.html")
    else:
        return redirect('/signin')

#------------------------------------------------------------------------------

def signin(request):
    if request.user.is_authenticated:
        return redirect('/home')
    else:
        if request.method=="POST":
            id=request.POST['id']
            password=request.POST['password']
            user=authenticate(username=id, password=password)
            if user is not None:
                login(request,user)
                return redirect('/home')
            else:
                messages.error(request, "Invalid username or password")
                return redirect('/signin')
        else:
            return render(request, "login.html")
        
#------------------------------------------------------------------------------

def signout(request):
    logout(request)
    return redirect('/')

#------------------------------------------------------------------------------

def signup(request):
    if request.method=="POST":
        cursor = connection.cursor()
        id = request.POST['id']
        name = request.POST['name']
        dept = request.POST['dept']
        prio = request.POST['prio']
        password = request.POST['password']
        cp = request.POST['cp']
        if password==cp:
            try:
                user = User.objects.create_user(username=id, first_name=name, password=password)
                user.save()
                cursor.execute("insert into faculty values ('{}', '{}', '{}', '{}')".format(
                    int(id), name, int(dept), int(prio)))
                connection.commit()
            except:
                messages.error(request, "Invalid entry.Couldn't add Faculty.")
            return redirect('/home/')
        else:
            return redirect('/signup')
    else:
        return render(request, "signup.html")
    
#------------------------------------------------------------------------------

def slot_pref(request):
    global iq, s1, s2, s3
    if request.user.is_authenticated:
        if request.method=="POST":
            m=sql.connect(host="localhost", user="root", passwd="1682896", database='tibblestemp')
            cursor=m.cursor()
            k=request.POST
            iq=request.user
            day = 'MONDAY'
            for key, value in k.items():
                if key=="mon1" and value != '0':
                    s1 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s1)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")

                if key=="mon2" and value != '0':
                    s2 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s2)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")
                if key=="mon3" and value != '0':
                    s3 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s3)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")
            day = 'TUESDAY'
            for key, value in k.items():
                if key=="tues1" and value != '0':
                    s1 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s1)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")
                if key=="tues2" and value != '0':
                    s2 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s2)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")
                if key=="tues3" and value != '0':
                    s3 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s3)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")
            day = 'WEDNESDAY'
            for key, value in k.items():
                if key=="wed1" and value != '0':
                    s1 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s1)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")
                if key=="wed2" and value != '0':
                    s2 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s2)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")
                if key=="wed3" and value != '0':
                    s3 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s3)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")
            day = 'THURSDAY'
            for key, value in k.items():
                if key=="thurs1" and value != '0':
                    s1 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s1)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")
                if key=="thurs2" and value != '0':
                    s2 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s2)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")
                if key=="thurs3" and value != '0':
                    s3 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s3)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")
            day = 'FRIDAY'
            for key, value in k.items():
                if key=="fri1" and value != '0':
                    s1 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s1)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")
                if key=="fri2" and value != '0':
                    s2 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s2)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")
                if key=="fri3" and value != '0':
                    s3 = value
                    c="insert into pref_slot values ('{}', '{}', '{}')".format(iq, day, s3)
                    try:
                        cursor.execute(c)
                        m.commit()
                    except:
                        messages.error(request, "Error submitting preferences")
            return redirect('/home')
        return render(request,'prefslot.html')
    else:
        return redirect('/')

#------------------------------------------------------------------------------
            
def sub_pref(request):
    global i, p1, p2, p3
    if request.user.is_authenticated:
        if request.method=="POST":
            m=sql.connect(host="localhost", user="root", passwd="1682896", database='tibblestemp')
            cursor=m.cursor()
            k=request.POST
            i=request.user
            for key, value in k.items():
                if key=="pref1":
                    p1 = value
                if key=="pref2":
                    p2 = value
                if key=="pref3":
                    p3 = value
            
            c="insert into pref_sub values ('{}', '{}', '{}', '{}')".format(i, p1, p2, p3)
            try:
                cursor.execute(c)
                m.commit()
                messages.success(request, "Preferences successfully submitted")
                return redirect('/prefsub/')
            except:
                messages.error(request, "Error submitting preferences")
                return redirect('/prefsub/')
        return render(request, 'prefsub.html')
    else:
        messages.error(request, "You must be logged in to submit preferences")
        return redirect('/')
    
#-----------------------------------------------------------------------

def display(request):
    cursor = connection.cursor()
    try:
        cursor.execute("TRUNCATE TABLE subject_allocation")
        cursor.execute("TRUNCATE TABLE slot_allocation_sem3")
        cursor.execute("TRUNCATE TABLE slot_allocation_sem5")
        cursor.execute("TRUNCATE TABLE timetable3")
        cursor.execute("TRUNCATE TABLE timetable5")
        cursor.execute("INSERT INTO timetable3 (DAY, PERIOD1, PERIOD2, PERIOD3, PERIOD4, PERIOD5, PERIOD6, PERIOD7) VALUES ('MONDAY', NULL, NULL, NULL, NULL, NULL, NULL, NULL), ('TUESDAY', NULL, NULL, NULL, NULL, NULL, NULL, NULL),  ('WEDNESDAY', NULL, NULL, NULL, NULL, NULL, NULL, NULL),  ('THURSDAY', NULL, NULL, NULL, NULL, NULL, NULL, NULL),  ('FRIDAY', NULL, NULL, NULL, NULL, NULL, NULL, NULL) ")
        cursor.execute("INSERT INTO timetable5 (DAY, PERIOD1, PERIOD2, PERIOD3, PERIOD4, PERIOD5, PERIOD6, PERIOD7) VALUES ('MONDAY', NULL, NULL, NULL, NULL, NULL, NULL, NULL), ('TUESDAY', NULL, NULL, NULL, NULL, NULL, NULL, NULL),  ('WEDNESDAY', NULL, NULL, NULL, NULL, NULL, NULL, NULL),  ('THURSDAY', NULL, NULL, NULL, NULL, NULL, NULL, NULL),  ('FRIDAY', NULL, NULL, NULL, NULL, NULL, NULL, NULL) ")
        cursor.callproc('pref1')
        cursor.callproc('pref2')
        cursor.callproc('pref3')
        for i in range (10):
            cursor.execute("CALL `remain`();")
            connection.commit()
        cursor.execute("select a.faculty_id, a.subject_code, s.hours_per_week, f.priority \
                        from subject s natural join subject_allocation a natural join   \
                        faculty f where s.semester = 3 and s.d_no = 1 order by f.priority, f.faculty_id")
        sem3_info = cursor.fetchall()
        cursor.execute("select distinct a.faculty_id, p.day, p.slot_no from subject_allocation a \
                        natural join pref_slot p where a.subject_code like '%3_'")
        sem3_slot = cursor.fetchall()
        tt3 = allot.allot3(sem3_info, sem3_slot)
        cursor.execute("select * from slot_allocation_sem3")
        sem3_tt = cursor.fetchall()
        if len(sem3_tt) == 0:
            for key in tt3.keys():
                if tt3[key] is not None:
                    query = "INSERT INTO slot_allocation_sem3 (faculty_id, subject_code, day, slot_no) VALUES (%s, %s, %s, %s)"
                    value = (tt3[key][0], tt3[key][1], key[0], key[1])
                    cursor.execute(query, value)
            cursor.callproc('tt3')
            cursor.execute("SELECT * FROM timetable3")
            timetable3 = cursor.fetchall()
        else:
            cursor.execute("SELECT * FROM timetable3")
            timetable3 = cursor.fetchall()
        connection.commit()
        cursor.execute("select a.faculty_id, a.subject_code, s.hours_per_week, f.priority \
                        from subject s natural join subject_allocation a natural join   \
                        faculty f where s.semester = 5 and s.d_no = 1 order by f.priority, f.faculty_id")
        sem5_info = cursor.fetchall()
        cursor.execute("select distinct a.faculty_id, p.day, p.slot_no from subject_allocation a \
                        natural join pref_slot p where a.subject_code like '%5_'")
        sem5_slot = cursor.fetchall()
        tt5 = allot.allot5(sem5_info, sem5_slot, tt3)
        cursor.execute("select * from slot_allocation_sem5")
        sem5_tt = cursor.fetchall()
        if len(sem5_tt) == 0:
            for key in tt5.keys():
                if tt5[key] is not None:
                    query = "INSERT INTO slot_allocation_sem5 (faculty_id, subject_code, day, slot_no) VALUES (%s, %s, %s, %s)"
                    value = (tt5[key][0], tt5[key][1], key[0], key[1])
                    cursor.execute(query, value)
            
            cursor.callproc('tt5')
            cursor.execute("SELECT * FROM timetable5")
            timetable5 = cursor.fetchall()
        else:
            cursor.execute("SELECT * FROM timetable5")
            timetable5 = cursor.fetchall()
        cursor.execute("SELECT s.subject_name, s.subject_code, f.faculty_name FROM subject_allocation a NATURAL JOIN faculty f NATURAL JOIN subject s WHERE subject_code like '%3_' ORDER BY s.subject_code ")
        faculty3 = cursor.fetchall()
        cursor.execute("SELECT s.subject_name, s.subject_code, f.faculty_name FROM subject_allocation a NATURAL JOIN faculty f NATURAL JOIN subject s WHERE subject_code like '%5_' ORDER BY s.subject_code ")
        faculty5 = cursor.fetchall()
        return render(request, 'timetable.html', {'displaydata3': timetable3, 'displaydata5': timetable5, 'displaydata7': faculty3, 'displaydata9': faculty5})
    except:
        messages.error(request, 'An error occurred. Check if all faculy members have given their preferences.')
        return redirect('/home')
    
#--------------------------------------------------------------------------------------

class GeneratePdf(View):
    def get(self, request,  *args, **kwargs):
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM timetable3")
        timetable3 = cursor.fetchall()
        cursor.execute("SELECT * FROM timetable5")
        timetable5 = cursor.fetchall()
        cursor.execute("SELECT s.subject_name, s.subject_code, f.faculty_name FROM subject_allocation a NATURAL JOIN faculty f NATURAL JOIN subject s WHERE subject_code like '%3_' ORDER BY s.subject_code ")
        faculty3 = cursor.fetchall()
        cursor.execute("SELECT s.subject_name, s.subject_code, f.faculty_name FROM subject_allocation a NATURAL JOIN faculty f NATURAL JOIN subject s WHERE subject_code like '%5_' ORDER BY s.subject_code ")
        faculty5 = cursor.fetchall()
        current_datetime = datetime.datetime.now()
        
        params = {
            'datetime': current_datetime,
            'displaydata3': timetable3,
            'displaydata5': timetable5,
            'displaydata7' : faculty3,
            'displaydata9' : faculty5
        }
        pdf = render_to_pdf('pdf.html', params)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "timetable.pdf"
            content = "inline; filename= %s" %(filename)
            response['Content-Disposition']=content
            return response
        return HttpResponse("Page Not Found")

#-----------------------------------------------------------------------------

def export_to_excel(request):
    workbook = Workbook()

    cursor = connection.cursor()
    cursor.execute("SELECT * FROM timetable3")
    timetable3 = cursor.fetchall()
    
    worksheet1 = workbook.active
    worksheet1.title = "3RD SEM TIMETABLE"
    
    for col_num, column_title in enumerate([i[0] for i in cursor.description]):
        worksheet1.cell(row=1,column=col_num+1).value = column_title
    for row_num, row_data in enumerate(timetable3, 2):
        for col_num, cell_data in enumerate(row_data):
            worksheet1.cell(row=row_num, column=col_num+1, value=cell_data)
    
    cursor.execute("SELECT * FROM timetable5")
    timetable5 = cursor.fetchall()
    
    worksheet2 = workbook.create_sheet(title="5TH SEM TIMETABLE")
    
    for col_num, column_title in enumerate([i[0] for i in cursor.description]):
        worksheet2.cell(row=1,column=col_num+1).value = column_title
    for row_num, row_data in enumerate(timetable5, 2):
        for col_num, cell_data in enumerate(row_data):
            worksheet2.cell(row=row_num, column=col_num+1, value=cell_data)
    
    cursor.execute("SELECT s.subject_name, s.subject_code, f.faculty_name FROM subject_allocation a NATURAL JOIN faculty f NATURAL JOIN subject s WHERE subject_code like '%3_' ORDER BY s.subject_code ")
    faculty3 = cursor.fetchall()
    
    worksheet3 = workbook.create_sheet(title="3RD SEM FACULTY LIST")
    
    for col_num, column_title in enumerate([i[0] for i in cursor.description]):
        worksheet3.cell(row=1,column=col_num+1).value = column_title
    for row_num, row_data in enumerate(faculty3, 2):
        for col_num, cell_data in enumerate(row_data):
            worksheet3.cell(row=row_num, column=col_num+1, value=cell_data)
    
    cursor.execute("SELECT s.subject_name, s.subject_code, f.faculty_name FROM subject_allocation a NATURAL JOIN faculty f NATURAL JOIN subject s WHERE subject_code like '%5_' ORDER BY s.subject_code ")
    faculty5 = cursor.fetchall()
    
    worksheet4 = workbook.create_sheet(title="5TH SEM FACULTY LIST")
    
    for col_num, column_title in enumerate([i[0] for i in cursor.description]):
        worksheet4.cell(row=1,column=col_num+1).value = column_title
    for row_num, row_data in enumerate(faculty5, 2):
        for col_num, cell_data in enumerate(row_data):
            worksheet4.cell(row=row_num, column=col_num+1, value=cell_data)
    
    first_row = worksheet1[1]
    for cell in first_row:
        cell.font = Font(name='Times New Roman', size=12, color='141414', bold=True)
        cell.fill = PatternFill(start_color="c7e8ff", end_color="c7e8ff", fill_type = "solid")
        cell.border = Border(left=Side(border_style='medium', color='6facd6'),
                       right=Side(border_style='medium', color='6facd6'),
                       top=Side(border_style='medium', color='6facd6'),
                       bottom=Side(border_style='medium', color='6facd6'))
        cell.alignment = Alignment(horizontal='center', vertical='center')

    worksheet1.column_dimensions["A"].width = 15
    worksheet1.column_dimensions["B"].width = 15
    worksheet1.column_dimensions["C"].width = 20
    worksheet1.column_dimensions["D"].width = 20
    worksheet1.column_dimensions["E"].width = 20
    worksheet1.column_dimensions["F"].width = 20
    worksheet1.column_dimensions["G"].width = 20
    worksheet1.column_dimensions["H"].width = 20
    
    first_row = worksheet2[1]
    for cell in first_row:
        cell.font = Font(name='Times New Roman', size=12, color='141414', bold=True)
        cell.fill = PatternFill(start_color="c7e8ff", end_color="c7e8ff", fill_type = "solid")
        cell.border = Border(left=Side(border_style='medium', color='6facd6'),
                       right=Side(border_style='medium', color='6facd6'),
                       top=Side(border_style='medium', color='6facd6'),
                       bottom=Side(border_style='medium', color='6facd6'))
        cell.alignment = Alignment(horizontal='center', vertical='center')

    worksheet2.column_dimensions["A"].width = 15
    worksheet2.column_dimensions["B"].width = 15
    worksheet2.column_dimensions["C"].width = 20
    worksheet2.column_dimensions["D"].width = 20
    worksheet2.column_dimensions["E"].width = 20
    worksheet2.column_dimensions["F"].width = 20
    worksheet2.column_dimensions["G"].width = 20
    worksheet2.column_dimensions["H"].width = 20
    
    first_row = worksheet3[1]
    for cell in first_row:
        cell.font = Font(name='Times New Roman', size=12, color='141414', bold=True)
        cell.fill = PatternFill(start_color="c7e8ff", end_color="c7e8ff", fill_type = "solid")
        cell.border = Border(left=Side(border_style='medium', color='6facd6'),
                       right=Side(border_style='medium', color='6facd6'),
                       top=Side(border_style='medium', color='6facd6'),
                       bottom=Side(border_style='medium', color='6facd6'))
        cell.alignment = Alignment(horizontal='center', vertical='center')

    worksheet3.column_dimensions["A"].width = 35
    worksheet3.column_dimensions["B"].width = 20
    worksheet3.column_dimensions["C"].width = 25
    
    first_row = worksheet4[1]
    for cell in first_row:
        cell.font = Font(name='Times New Roman', size=12, color='141414', bold=True)
        cell.fill = PatternFill(start_color="c7e8ff", end_color="c7e8ff", fill_type = "solid")
        cell.border = Border(left=Side(border_style='medium', color='6facd6'),
                       right=Side(border_style='medium', color='6facd6'),
                       top=Side(border_style='medium', color='6facd6'),
                       bottom=Side(border_style='medium', color='6facd6'))
        cell.alignment = Alignment(horizontal='center', vertical='center')

    worksheet4.column_dimensions["A"].width = 35
    worksheet4.column_dimensions["B"].width = 20
    worksheet4.column_dimensions["C"].width = 25
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=aiml.xlsx'
    workbook.save(response)
    return response

# -----------------------------------------------------------------------------------

def subp(request):
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM pref_sub")
    fac = cursor.fetchall()
    return render(request, 'subp.html', {'fact': fac})

# -----------------------------------------------------------------------------------

def slotp(request):
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM pref_slot")
    fac = cursor.fetchall()
    return render(request, 'slotp.html', {'fact': fac})

#-----------------------------------------------------------------------------------

def viewfac(request):
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM faculty WHERE D_NO=1")
    fac = cursor.fetchall()
    return render(request, 'faculty.html', {'fact': fac})

# -----------------------------------------------------------------------------------

def delfac(request, faculty_id):
    if request.user.is_authenticated:
        cursor = connection.cursor()
        try:
            cursor.execute("DELETE FROM faculty WHERE faculty_id='{}'".format(faculty_id))
            connection.commit()
            messages.success(request, 'You have deleted Faculty')
        except IntegrityError as e:
            messages.error(
                request, 'An error occurred while deleting the entry: {}'.format(e))
            return redirect('home.html')
        except Exception as e:
            messages.error(
                request, 'An error occurred while deleting the entry: {}'.format(e))
            return redirect('home.html')
        return render(request, 'home.html')
    
#------------------------------------------------------------------------------------

def viewsub(request):
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM subject WHERE D_NO=1")
    sub = cursor.fetchall()
    return render(request, 'subject.html', {'subj': sub})

# --------------------------------------------------------------------------------------

def addsub(request):
    if request.user.is_authenticated:
        if request.method == "POST":
            m = sql.connect(host="localhost", user="root",
                            passwd="1682896", database='tibblestemp')
            cursor = m.cursor()
            k = request.POST
            for key, value in k.items():
                if key == "subject_code":
                    p1 = value
                if key == "dno":
                    p2 = value
                if key == "subject_name":
                    p3 = value
                if key == "semester":
                    p4 = value
                if key == "hpw":
                    p5 = value
            c = "insert into subject values ('{}', '{}', '{}', '{}', '{}')".format(p1, p2, p3, p4, p5)
            try:
                cursor.execute(c)
                m.commit()
                messages.success(request, "Subject successfully added.")
                return redirect('/home/')
            except:
                messages.error(request, "Error adding Subject.")
                return redirect('/home/')
        return render(request, 'addsub.html')
    else:
        messages.error(request, "You must be logged in to submit preferences")
        return redirect('/')

# --------------------------------------------------------------------------------------

def delsub(request, subject_code):
    if request.user.is_authenticated:
        cursor = connection.cursor()
        try:
            cursor.execute("DELETE FROM subject WHERE subject_code='{}'".format(subject_code))
            connection.commit()
            messages.success(request, 'You have deleted the Subject')
        except IntegrityError as e:
            messages.error(
                request, 'An error occurred while deleting the entry: {}'.format(e))
            return redirect('home.html')
        except Exception as e:
            messages.error(request, 'An error occurred while deleting the entry: {}'.format(e))
            return redirect('home.html')
        return render(request, 'home.html')

#--------------------------------------------------------------------------------------

def view_slotentry(request):
    if request.user.is_authenticated:
        id = request.user
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM pref_slot WHERE faculty_id=('{}')".format(id))
        ps = cursor.fetchall()
        return render(request, 'slotsview.html', {'pss': ps})

# --------------------------------------------------------------------------------------

def delete_slotentry(request, faculty_id, day, slot):
    if request.user.is_authenticated:
        cursor = connection.cursor()
        try:
            cursor.execute("DELETE FROM pref_slot WHERE faculty_id=('{}') AND day=('{}') AND slot_no = ('{}')".format(faculty_id, day, slot))
            connection.commit()
            messages.success(request, 'Your Slot Preference is deleted')
        except IntegrityError as e:
            messages.error(request, 'An error occurred while deleting the entry: {}'.format(e))
            return redirect('prefslot.html')
        except Exception as e:
            messages.error(request, 'An error occurred while deleting the entry: {}'.format(e))
            return redirect('prefslot.html')
        return render(request, 'prefslot.html')

# --------------------------------------------------------------------------------------

def view_subentry(request):
    if request.user.is_authenticated:
        id = request.user
        cursor = connection.cursor()
        cursor.execute(
            "SELECT * FROM pref_sub WHERE faculty_id=('{}')".format(id))
        ps = cursor.fetchall()
        return render(request, 'subsview.html', {'pss': ps})
    
# --------------------------------------------------------------------------------------

def delete_subentry(request):
    if request.user.is_authenticated:
        id = request.user
        cursor = connection.cursor()
        try:
            cursor.execute(
                "DELETE FROM pref_sub WHERE faculty_id=('{}')".format(id))
            messages.success(request, 'Your Preferences are deleted')
        except Exception as e:
            messages.error(
                request, 'An error occurred while deleting the entry: {}'.format(e))
            return redirect('error.html')
        return render(request, 'prefsub.html')

#---------------------------------------------------------------------------------------------

def updatesub(request):
    if request.user.is_authenticated:
        if request.method == "POST":
            m = sql.connect(host="localhost", user="root",
                            passwd="1682896", database='tibblestemp')
            cursor = m.cursor()
            k = request.POST
            i = request.user
            for key, value in k.items():
                if key == "pref1":
                    p1 = value
                if key == "pref2":
                    p2 = value
                if key == "pref3":
                    p3 = value
            c = cursor.execute("UPDATE pref_sub SET pref_1 = %s, pref_2 = %s, pref_3 = %s WHERE faculty_id = %s;", (p1, p2, p3, str(i)))
            cursor.execute(c)
            m.commit()
            return redirect('/viewsubpref/')
        return render(request,'updatesub.html')
    else:
        return redirect('/')

# ---------------------------------------------------------------------------------------------
